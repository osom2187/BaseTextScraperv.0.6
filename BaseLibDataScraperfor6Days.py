import openpyxl as op
import pandas as pd
import xlsxwriter as xw
import os
import glob

all_nec_files = glob.glob('C:\\Users\\dav\\statistik\\*.txt')

here_is_everything = []
for file in all_nec_files:
    txt = open(file, 'r+')
    here_is_everything.append(txt.read())

data1 = here_is_everything[1]
data2 = here_is_everything[2]
data3 = here_is_everything[3]
data4 = here_is_everything[4]
data5 = here_is_everything[5]
data6 = here_is_everything[6]
data4week = data1, data2, data3, data4, data5, data6

header = data1[61:106]
precioustable = data1[312:884].split()

preciousNums = []

for item in precioustable:
    try:
        preciousNums.append(int(item))
    except:
        preciousNums.append(item)
        continue

precioustable = preciousNums

table1_colNames = data1[162:234].split()  # 10 items in list
table1_rowNames = [data1[308:311], data1[380:384], data1[452:456], data1[524:528], data1[596:600], data1[668:676],
                   data1[740:743], data1[812:818]]  # 8 items
table1_nums1 = data1[312:379].split()
table1_nums2 = data1[385:453].split()
table1_nums3 = data1[457:523].split()
table1_nums4 = data1[529:595].split()
table1_nums5 = data1[601:667].split()
table1_nums6 = data1[677:739].split()
table1_nums7 = data1[744:811].split()
table1_nums8 = data1[819:885].split()

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'lets see'

sheet = wb['lets see']

# first round

ws['A1'] = header
ws['A3'] = table1_colNames[0]
ws['B3'] = table1_colNames[1]
ws['C3'] = table1_colNames[2]
ws['D3'] = table1_colNames[3]
ws['E3'] = table1_colNames[4]
ws['F3'] = table1_colNames[5]
ws['G3'] = table1_colNames[6]
ws['H3'] = table1_colNames[7]
ws['I3'] = table1_colNames[8]
ws['J3'] = table1_colNames[9]

post109a = precioustable[9:19]
post109b = precioustable[19:29]
post109h = precioustable[29:39]
post109j = precioustable[39:49]
postSonstige = precioustable[49:59]
postWeb = precioustable[59:69]
postGesamt = precioustable[69:79]
IsThisDf = post109a, post109b, post109h, post109j, postSonstige, postWeb, postGesamt
df = pd.DataFrame(data=IsThisDf)

from openpyxl.utils.dataframe import dataframe_to_rows

for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

# second round

header = data2[61:106]
precioustable = data2[312:884].split()

preciousNums = []

for item in precioustable:
    try:
        preciousNums.append(int(item))
    except:
        preciousNums.append(item)
        continue

precioustable = preciousNums

ws['A12'] = header
ws['A14'] = table1_colNames[0]
ws['B14'] = table1_colNames[1]
ws['C14'] = table1_colNames[2]
ws['D14'] = table1_colNames[3]
ws['E14'] = table1_colNames[4]
ws['F14'] = table1_colNames[5]
ws['G14'] = table1_colNames[6]
ws['H14'] = table1_colNames[7]
ws['I14'] = table1_colNames[8]
ws['J14'] = table1_colNames[9]

post109a = precioustable[9:19]
post109b = precioustable[19:29]
post109h = precioustable[29:39]
post109j = precioustable[39:49]
postSonstige = precioustable[49:59]
postWeb = precioustable[59:69]
postGesamt = precioustable[69:79]
IsThisDf = post109a, post109b, post109h, post109j, postSonstige, postWeb, postGesamt
df = pd.DataFrame(data=IsThisDf)

for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

# third round

header = data3[61:106]
precioustable = data3[312:884].split()

preciousNums = []

for item in precioustable:
    try:
        preciousNums.append(int(item))
    except:
        preciousNums.append(item)
        continue

precioustable = preciousNums

ws['A23'] = header
ws['A25'] = table1_colNames[0]
ws['B25'] = table1_colNames[1]
ws['C25'] = table1_colNames[2]
ws['D25'] = table1_colNames[3]
ws['E25'] = table1_colNames[4]
ws['F25'] = table1_colNames[5]
ws['G25'] = table1_colNames[6]
ws['H25'] = table1_colNames[7]
ws['I25'] = table1_colNames[8]
ws['J25'] = table1_colNames[9]

post109a = precioustable[9:19]
post109b = precioustable[19:29]
post109h = precioustable[29:39]
post109j = precioustable[39:49]
postSonstige = precioustable[49:59]
postWeb = precioustable[59:69]
postGesamt = precioustable[69:79]
IsThisDf = post109a, post109b, post109h, post109j, postSonstige, postWeb, postGesamt
df = pd.DataFrame(data=IsThisDf)

for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

# fourth round

header = data4[61:106]
precioustable = data4[312:884].split()

preciousNums = []

for item in precioustable:
    try:
        preciousNums.append(int(item))
    except:
        preciousNums.append(item)
        continue

precioustable = preciousNums

ws['A34'] = header
ws['A36'] = table1_colNames[0]
ws['B36'] = table1_colNames[1]
ws['C36'] = table1_colNames[2]
ws['D36'] = table1_colNames[3]
ws['E36'] = table1_colNames[4]
ws['F36'] = table1_colNames[5]
ws['G36'] = table1_colNames[6]
ws['H36'] = table1_colNames[7]
ws['I36'] = table1_colNames[8]
ws['J36'] = table1_colNames[9]

post109a = precioustable[9:19]
post109b = precioustable[19:29]
post109h = precioustable[29:39]
post109j = precioustable[39:49]
postSonstige = precioustable[49:59]
postWeb = precioustable[59:69]
postGesamt = precioustable[69:79]
IsThisDf = post109a, post109b, post109h, post109j, postSonstige, postWeb, postGesamt
df = pd.DataFrame(data=IsThisDf)

for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

# fifth round

header = data5[61:106]
precioustable = data5[312:884].split()

preciousNums = []

for item in precioustable:
    try:
        preciousNums.append(int(item))
    except:
        preciousNums.append(item)
        continue

precioustable = preciousNums

ws['A45'] = header
ws['A47'] = table1_colNames[0]
ws['B47'] = table1_colNames[1]
ws['C47'] = table1_colNames[2]
ws['D47'] = table1_colNames[3]
ws['E47'] = table1_colNames[4]
ws['F47'] = table1_colNames[5]
ws['G47'] = table1_colNames[6]
ws['H47'] = table1_colNames[7]
ws['I47'] = table1_colNames[8]
ws['J47'] = table1_colNames[9]

post109a = precioustable[9:19]
post109b = precioustable[19:29]
post109h = precioustable[29:39]
post109j = precioustable[39:49]
postSonstige = precioustable[49:59]
postWeb = precioustable[59:69]
postGesamt = precioustable[69:79]
IsThisDf = post109a, post109b, post109h, post109j, postSonstige, postWeb, postGesamt
df = pd.DataFrame(data=IsThisDf)

for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

# sixth round

header = data6[61:106]
precioustable = data6[312:884].split()

preciousNums = []

for item in precioustable:
    try:
        preciousNums.append(int(item))
    except:
        preciousNums.append(item)
        continue

precioustable = preciousNums

ws['A56'] = header
ws['A58'] = table1_colNames[0]
ws['B58'] = table1_colNames[1]
ws['C58'] = table1_colNames[2]
ws['D58'] = table1_colNames[3]
ws['E58'] = table1_colNames[4]
ws['F58'] = table1_colNames[5]
ws['G58'] = table1_colNames[6]
ws['H58'] = table1_colNames[7]
ws['I58'] = table1_colNames[8]
ws['J58'] = table1_colNames[9]

post109a = precioustable[9:19]
post109b = precioustable[19:29]
post109h = precioustable[29:39]
post109j = precioustable[39:49]
postSonstige = precioustable[49:59]
postWeb = precioustable[59:69]
postGesamt = precioustable[69:79]
IsThisDf = post109a, post109b, post109h, post109j, postSonstige, postWeb, postGesamt

df = pd.DataFrame(data=IsThisDf)

for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

wb.save('lets_see.xlsx')
